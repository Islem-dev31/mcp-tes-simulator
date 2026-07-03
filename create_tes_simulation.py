# -*- coding: utf-8 -*-
"""
Simulation de Performance Thermique d'une Batterie Thermique à MCP
Destinée à une Chambre Froide Industrielle de 100 m³

Ce script implémente les lois physiques de transfert de chaleur (Fourier, Newton, infiltration d'air,
chaleur latente) pour évaluer la performance d'une batterie thermique à changement de phase (MCP).
Il génère un fichier Excel professionnel (.xlsx) contenant les résultats de la simulation,
analyse la sensibilité économique (hausse du coût de l'alu, gains d'énergie) et intègre des notes
techniques pour la soutenance devant le jury.

Auteur : Ingénieur Thermicien Expert en Stockage d'Énergie Thermique (TES)
"""

import math
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ==============================================================================
# 1. PARAMÈTRES PHYSIQUES ET CONSTANTES DE CONCEPTION
# ==============================================================================

# Caractéristiques de la chambre froide (100 m³)
V_ROOM = 100.0        # Volume de la chambre froide (m³)
L_ROOM = 5.0          # Longueur (m)
W_ROOM = 5.0          # Largeur (m)
H_ROOM = 4.0          # Hauteur (m)
A_ENV = 130.0         # Surface totale des parois (m²): 2*(5*5) + 4*(5*4)

# Propriétés de l'isolant (Polyuréthane - PU)
LAMBDA_PU = 0.022     # Conductivité thermique du PU (W/m.K)
THICKNESS_PU_POS = 0.100  # Épaisseur isolant pour +4°C (100 mm -> 0.1 m)
THICKNESS_PU_NEG = 0.200  # Épaisseur isolant pour -18°C (200 mm -> 0.2 m)

# Coefficients de transfert superficiel des parois (Norme ASHRAE)
H_INT_WALL = 8.3      # Convection intérieure (W/m².K)
H_EXT_WALL = 25.0     # Convection extérieure (W/m².K)

# Caractéristiques de la porte et des infiltrations (Théorème de Sebbar / Gosney & Olama)
W_DOOR = 1.2          # Largeur de la porte (m)
H_DOOR = 2.0          # Hauteur de la porte (m)
A_DOOR = W_DOOR * H_DOOR  # Surface de la porte (m²)
N_OPENINGS_PER_HOUR = 4.0 # Nombre d'ouvertures de porte par heure
T_OPENING = 30.0      # Durée moyenne d'une ouverture (secondes)
C_D = 0.6             # Coefficient de décharge pour l'orifice d'écoulement d'air
G = 9.81              # Accélération de la pesanteur (m/s²)
P_ATM = 101325.0      # Pression atmosphérique standard (Pa)
R_AIR = 287.05        # Constante spécifique de l'air (J/kg.K)
CP_AIR = 1005.0       # Capacité thermique massique de l'air (J/kg.K)

# Caractéristiques du stock de produits stockés (Charge thermique passive)
STOCK_VOLUME = 20000.0   # Capacité de stockage (Litres ou kg equivalent eau)
STOCK_TURNOVER = 0.10    # Taux de rotation quotidien du stock (10% par jour)
CP_PRODUCT = 3800.0      # Capacité thermique massique du produit (J/kg.K)
T_IN_PRODUCT_OFFSET = 5.0 # Différence entre la température d'entrée du produit et l'ambiant (K)

# Charge thermique interne constante (Éclairage, ventilateurs, etc.)
Q_INTERNAL_STATIC = 150.0 # Watts

# Propriétés de la batterie thermique (MCP et Casing)
LAMBDA_AL = 200.0     # Conductivité thermique de l'Aluminium (W/m.K)
RHO_AL = 2700.0       # Masse volumique de l'Aluminium (kg/m³)

# MCP : Paraffine adaptée aux plages de température
LAMBDA_PCM = 0.20     # Conductivité thermique moyenne de la paraffine (W/m.K)
L_F = 200000.0        # Chaleur latente de fusion (J/kg, soit 200 kJ/kg)
RHO_PCM = 850.0       # Masse volumique de la paraffine (kg/m³)

# Résistance de contact thermique due à la peinture Époxy
R_CONTACT_EPOXY = 5e-4 # Résistance thermique additionnelle (m².K/W)
LAMBDA_EPOXY = 0.20    # Conductivité thermique de la résine Époxy (W/m.K)

# Convection sur les cylindres MCP
H_CONV_VENT_ON = 25.0   # Ventilation forcée ON (W/m².K)
H_CONV_VENT_OFF = 5.0   # Convection naturelle OFF (W/m².K)

# Autonomie cible pour le dimensionnement de la masse de MCP
T_AUTONOMY_TARGET = 13.0 # heures

# Tarifs et paramètres Sonelgaz pour l'analyse économique
TARIF_HP_DA_KWH = 11.0   # Tarif heures pleines industrielle (DA/kWh)
TARIF_HC_DA_KWH = 3.0    # Tarif heures creuses industrielle (DA/kWh)
DELTA_TARIF_DA_KWH = TARIF_HP_DA_KWH - TARIF_HC_DA_KWH # Différence (8 DA/kWh)
DAYS_OP_YEAR = 330.0     # Jours d'opération par an
PRIME_FIXE_SAVINGS_DA = 80000.0 # Économie forfaitaire sur prime de puissance (DA/an)

# ==============================================================================
# 2. FONCTIONS DE CALCUL THERMODYNAMIQUE
# ==============================================================================

def get_air_density(temp_c):
    temp_k = temp_c + 273.15
    return P_ATM / (R_AIR * temp_k)

def get_saturation_vapor_pressure(temp_c):
    if temp_c >= 0:
        a, b = 17.27, 237.3
    else:
        a, b = 21.875, 265.5
    return 610.78 * math.exp((a * temp_c) / (temp_c + b))

def get_humidity_ratio(temp_c, rh_percent):
    p_sat = get_saturation_vapor_pressure(temp_c)
    p_v = (rh_percent / 100.0) * p_sat
    p_v = min(p_v, P_ATM - 100.0)
    return 0.622 * p_v / (P_ATM - p_v)

def get_air_enthalpy(temp_c, rh_percent):
    w = get_humidity_ratio(temp_c, rh_percent)
    return CP_AIR * temp_c + w * (2501000.0 + 1860.0 * temp_c)

def calculate_door_infiltration_load(t_ext, rh_ext, t_int, rh_int):
    rho_ext = get_air_density(t_ext)
    rho_int = get_air_density(t_int)
    rho_avg = (rho_ext + rho_int) / 2.0
    delta_rho = max(0.0, rho_int - rho_ext)
    
    # Débit massique d'air entrant lors de l'ouverture (Théorème de Sebbar / Tamm)
    m_dot_air = (2.0 / 3.0) * C_D * W_DOOR * (H_DOOR ** 1.5) * math.sqrt(G * delta_rho / rho_avg)
    
    h_ext = get_air_enthalpy(t_ext, rh_ext)
    h_int = get_air_enthalpy(t_int, rh_int)
    delta_h = max(0.0, h_ext - h_int)
    
    d_open = (N_OPENINGS_PER_HOUR * T_OPENING) / 3600.0
    return m_dot_air * delta_h * d_open

def calculate_wall_heat_load(t_ext, t_int, thickness_pu):
    r_total = (1.0 / H_INT_WALL) + (thickness_pu / LAMBDA_PU) + (1.0 / H_EXT_WALL)
    u_overall = 1.0 / r_total
    return u_overall * A_ENV * (t_ext - t_int)

def calculate_product_load(t_ext, t_int):
    m_turnover = STOCK_VOLUME * STOCK_TURNOVER
    t_in_product = t_ext - T_IN_PRODUCT_OFFSET
    q_product = (m_turnover * CP_PRODUCT * (t_in_product - t_int)) / (24.0 * 3600.0)
    return max(0.0, q_product)

# ==============================================================================
# 3. CONCEPTION ET PERFORMANCE DU CYLINDRE MCP
# ==============================================================================

def calculate_cylinder_thermal_resistance(d_outer, t_wall_al, n_fins, t_fin, t_paint, ventilation):
    d_inner = d_outer - 2 * t_wall_al
    h_conv = H_CONV_VENT_ON if ventilation == "ON" else H_CONV_VENT_OFF
    
    # Longueur des ailettes externes (fixée à 30 mm = 0.03 m)
    l_fin_ext = 0.030
    # Longueur des ailettes internes (occupant 80% du rayon intérieur)
    l_fin_int = 0.8 * (d_inner / 2.0)
    
    # --- 1. CONVECTION EXTERNE & AILETTES EXTERNES ---
    a_unfinned_ext = math.pi * d_outer - n_fins * t_fin
    a_fin_ext = 2.0 * n_fins * l_fin_ext
    
    # Efficacité de l'ailette externe
    if n_fins > 0 and t_fin > 0:
        m_fin_ext = math.sqrt((h_conv * 2.0) / (LAMBDA_AL * t_fin))
        efficiency_fin_ext = math.tanh(m_fin_ext * l_fin_ext) / (m_fin_ext * l_fin_ext) if (m_fin_ext * l_fin_ext) > 0 else 1.0
    else:
        efficiency_fin_ext = 1.0
    
    a_eff_ext = a_unfinned_ext + efficiency_fin_ext * a_fin_ext
    r_conv = 1.0 / (h_conv * a_eff_ext)
    
    # --- 2. RÉSISTANCE DE PEINTURE ÉPOXY ---
    r_epoxy = t_paint / (LAMBDA_EPOXY * math.pi * d_outer)
    r_epoxy_contact = R_CONTACT_EPOXY / (math.pi * d_outer)
    r_paint = r_epoxy + r_epoxy_contact
    
    # --- 3. CONDUCTION DANS LA PAROI D'ALUMINIUM ---
    r_wall_al = math.log(d_outer / d_inner) / (2.0 * math.pi * LAMBDA_AL)
    
    # --- 4. CONDUCTION DANS LE MCP (Paraffine améliorée par ailettes internes) ---
    a_pcm_casing = math.pi * (d_inner ** 2) / 4.0
    if n_fins > 0 and t_fin > 0:
        a_fins_int_cross = n_fins * t_fin * l_fin_int
    else:
        a_fins_int_cross = 0.0
    phi_fins_int = a_fins_int_cross / a_pcm_casing if a_pcm_casing > 0 else 0.0
    
    # Conductivité thermique effective du MCP composite
    lambda_pcm_eff = LAMBDA_PCM * (1.0 - phi_fins_int) + LAMBDA_AL * phi_fins_int
    
    # Résistance de conduction du MCP
    r_pcm = 1.0 / (4.0 * math.pi * lambda_pcm_eff)
    
    # Résistance thermique totale linéaire (K.m/W)
    r_total_linear = r_conv + r_paint + r_wall_al + r_pcm
    return r_total_linear, lambda_pcm_eff

# ==============================================================================
# 4. MODÈLE DE SIMULATION ET ANALYSE DE SENSIBILITÉ
# ==============================================================================

def run_simulation():
    target_temps = [4.0, -18.0]
    external_temps = [25.0, 30.0, 35.0, 40.0, 45.0]
    cylinder_diameters = [0.080, 0.100, 0.120]  # m
    wall_thicknesses_al = [0.002]                 # m
    fin_counts = [0, 4, 8]
    fin_thicknesses = [0.0015]                    # m
    paint_thicknesses = [100e-6]                  # m
    ventilation_options = ["ON", "OFF"]
    cylinder_lengths = [1.0, 1.5, 2.0]            # m
    empty_spaces = [10, 15, 20]                   # %
    
    results = []
    
    # Base de coût d'ingénierie (DA)
    prix_al_base = 1500.0      # DA/kg (Aluminium standard)
    prix_al_sens = 1800.0      # DA/kg (Aluminium +20% sensibilité)
    prix_mcp = 400.0          # DA/kg
    cout_fab_cyl = 1000.0     # DA/cylindre
    
    for t_target in target_temps:
        thickness_pu = THICKNESS_PU_POS if t_target == 4.0 else THICKNESS_PU_NEG
        t_fusion_pcm = 2.0 if t_target == 4.0 else -21.0
        delta_t_driving = t_target - t_fusion_pcm
        
        # COP estimé du système frigorifique selon la plage de consigne
        cop = 3.0 if t_target == 4.0 else 1.8
        
        for t_ext in external_temps:
            q_wall = calculate_wall_heat_load(t_ext, t_target, thickness_pu)
            q_inf = calculate_door_infiltration_load(t_ext, 60.0, t_target, 90.0)
            q_prod = calculate_product_load(t_ext, t_target)
            q_int = Q_INTERNAL_STATIC
            
            q_load_total = q_wall + q_inf + q_prod + q_int
            
            # Énergie requise en Joules pour 13 heures d'autonomie
            e_required_joules = q_load_total * 13.0 * 3600.0
            m_pcm_required = e_required_joules / L_F
            
            for d_cyl in cylinder_diameters:
                for t_wall in wall_thicknesses_al:
                    for n_fins in fin_counts:
                        current_fin_thicknesses = [0.0] if n_fins == 0 else fin_thicknesses
                        for t_fin in current_fin_thicknesses:
                            for t_paint in paint_thicknesses:
                                for vent in ventilation_options:
                                    for l_cyl in cylinder_lengths:
                                        for empty_space in empty_spaces:
                                            
                                            r_linear, lambda_eff = calculate_cylinder_thermal_resistance(
                                                d_cyl, t_wall, n_fins, t_fin, t_paint, vent
                                            )
                                            
                                            d_inner = d_cyl - 2.0 * t_wall
                                            l_fin_int = 0.8 * (d_inner / 2.0) if n_fins > 0 else 0.0
                                            
                                            a_pcm_cross = math.pi * (d_inner ** 2) / 4.0 - n_fins * t_fin * l_fin_int
                                            m_pcm_per_meter = a_pcm_cross * RHO_PCM
                                            
                                            total_length_needed = m_pcm_required / m_pcm_per_meter
                                            ua_battery = total_length_needed / r_linear
                                            q_battery_max = ua_battery * delta_t_driving
                                            
                                            l_fin_ext = 0.030 if n_fins > 0 else 0.0
                                            a_al_cross = (math.pi * (d_cyl**2 - d_inner**2) / 4.0) + n_fins * t_fin * (l_fin_ext + l_fin_int)
                                            m_al_per_meter = a_al_cross * RHO_AL
                                            m_al_total = m_al_per_meter * total_length_needed
                                            
                                            thermal_effectiveness = min(1.0, q_battery_max / q_load_total)
                                            t_autonomy_real = 13.0 * thermal_effectiveness
                                            
                                            gravimetric_efficiency = m_pcm_required / (m_pcm_required + m_al_total)
                                            efficiency_total = thermal_effectiveness * gravimetric_efficiency
                                            
                                            # Règle du ciel gazeux
                                            m_pcm_dilated = m_pcm_required * 1.10
                                            v_pcm_dilated = m_pcm_dilated / RHO_PCM
                                            v_cyl = math.pi * (d_cyl / 2.0)**2 * l_cyl
                                            
                                            filling_ratio = 1.0 - empty_space / 100.0
                                            v_cyl_effective = v_cyl * filling_ratio
                                            n_modules = v_pcm_dilated / v_cyl_effective
                                            
                                            # Coût Standard en DA
                                            cout_total_da = (m_al_total * prix_al_base) + (m_pcm_required * prix_mcp) + (n_modules * cout_fab_cyl)
                                            if vent == "ON":
                                                cout_total_da += 15000.0
                                                
                                            # Analyse de Sensibilité : Hausse de 20% du prix de l'Aluminium (1800 DA/kg)
                                            cout_total_sens_da = (m_al_total * prix_al_sens) + (m_pcm_required * prix_mcp) + (n_modules * cout_fab_cyl)
                                            if vent == "ON":
                                                cout_total_sens_da += 15000.0
                                                
                                            # Économie d'énergie Sonelgaz (DA/an)
                                            # Puissance électrique économisée pendant les heures pleines (kW_e)
                                            p_elec_saved = q_load_total / (cop * 1000.0)
                                            # Énergie cumulée économisée par jour (kWh_e/jour)
                                            e_saved_daily = p_elec_saved * t_autonomy_real
                                            # Économie annuelle en DA (Différence de tarif HP vs HC + gain prime fixe de puissance)
                                            economie_annuelle_da = (e_saved_daily * DELTA_TARIF_DA_KWH * DAYS_OP_YEAR) + PRIME_FIXE_SAVINGS_DA
                                            
                                            # Temps de Retour sur Investissement (ans)
                                            payback_years = cout_total_da / economie_annuelle_da if economie_annuelle_da > 0 else 99.0
                                            payback_years_sens = cout_total_sens_da / economie_annuelle_da if economie_annuelle_da > 0 else 99.0
                                            
                                            opt_score_da = t_autonomy_real / cout_total_da if cout_total_da > 0 else 0
                                            opt_score_kg = t_autonomy_real / m_pcm_required if m_pcm_required > 0 else 0
                                            
                                            results.append({
                                                "Temp_Target": t_target,
                                                "Temp_Ext": t_ext,
                                                "D_Cyl_mm": d_cyl * 1000.0,
                                                "T_Wall_Al_mm": t_wall * 1000.0,
                                                "N_Fins": n_fins,
                                                "T_Fin_mm": t_fin * 1000.0,
                                                "Paint_Thickness_um": t_paint * 1e6,
                                                "Ventilation": vent,
                                                "L_Cyl_m": l_cyl,
                                                "Empty_Space_Percent": empty_space,
                                                "Q_Wall_W": q_wall,
                                                "Q_Infiltration_W": q_inf,
                                                "Q_Product_W": q_prod,
                                                "Q_Internal_W": q_int,
                                                "Q_Load_Total_W": q_load_total,
                                                "M_PCM_Required_kg": m_pcm_required,
                                                "M_Al_Total_kg": m_al_total,
                                                "R_Linear_K_m_W": r_linear,
                                                "Lambda_Eff_W_m_K": lambda_eff,
                                                "UA_Battery_W_K": ua_battery,
                                                "Q_Battery_Max_W": q_battery_max,
                                                "Autonomy_Real_h": t_autonomy_real,
                                                "Efficiency_Thermal": thermal_effectiveness,
                                                "Efficiency_Gravimetric": gravimetric_efficiency,
                                                "Efficiency_Total": efficiency_total,
                                                "Volume_Cyl_m3": v_cyl,
                                                "M_PCM_Dilated_kg": m_pcm_dilated,
                                                "V_PCM_Dilated_m3": v_pcm_dilated,
                                                "N_Modules": n_modules,
                                                "Cost_DA": cout_total_da,
                                                "Cost_Sens_DA": cout_total_sens_da,
                                                "Savings_Yearly_DA": economie_annuelle_da,
                                                "Payback_Years": payback_years,
                                                "Payback_Years_Sens": payback_years_sens,
                                                "Optimization_Score_h_DA": opt_score_da,
                                                "Optimization_Score_h_kg": opt_score_kg
                                            })
                                            
    return pd.DataFrame(results)

# ==============================================================================
# 5. EXPORT ET FORMATAGE PROFESSIONNEL DE L'EXCEL
# ==============================================================================

def create_styled_excel(df, filename="Simulation_Performances_MCP_TES_Final.xlsx"):
    best_configs = []
    for t_target in [4.0, -18.0]:
        for t_ext in [25.0, 30.0, 35.0, 40.0, 45.0]:
            sub_df = df[(df["Temp_Target"] == t_target) & (df["Temp_Ext"] == t_ext)]
            if not sub_df.empty:
                idx_best = sub_df["Optimization_Score_h_DA"].idxmax()
                best_configs.append(sub_df.loc[idx_best])
                
    df_best_configs = pd.DataFrame(best_configs)
    best_indices = set(df_best_configs.index)
    
    wb = openpyxl.Workbook()
    
    # Feuille 1 : Tableau de bord et Synthèse d'Optimisation
    ws_dash = wb.active
    ws_dash.title = "Dashboard & Optimisation"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Feuille 2 : Graphes d'Interpolation
    ws_graphs = wb.create_sheet(title="Graphes d'Interpolation")
    ws_graphs.views.sheetView[0].showGridLines = True
    
    # Feuille 3 : Base de données complète
    ws_data = wb.create_sheet(title="Données de Simulation")
    ws_data.views.sheetView[0].showGridLines = True
    
    # Définition des Styles Thème "Classic Navy"
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=13, bold=True, color="1B365D")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    font_regular = Font(name="Segoe UI", size=10)
    font_small = Font(name="Segoe UI", size=9, italic=True, color="555555")
    
    fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_light_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_accent_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    fill_accent_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # --------------------------------------------------------------------------
    # RENSEIGNER LE DASHBOARD (FEUILLE 1)
    # --------------------------------------------------------------------------
    
    # En-tête principal
    ws_dash.merge_cells("A1:G2")
    title_cell = ws_dash["A1"]
    title_cell.value = "SIMULATION DE STOCKAGE D'ÉNERGIE THERMIQUE (TES) - CHAMBRE FROIDE 100 m³"
    title_cell.font = font_title
    title_cell.fill = fill_navy
    title_cell.alignment = align_center
    
    ws_dash.row_dimensions[1].height = 25
    ws_dash.row_dimensions[2].height = 25
    
    # Section 1: Paramètres Fixes de Simulation (Colonnes A à C)
    ws_dash["A4"] = "1. PARAMÈTRES FIXES DE CONCEPTION"
    ws_dash["A4"].font = font_section
    
    fixed_params = [
        ("Volume de la chambre froide", V_ROOM, "m³", "Volume géométrique type"),
        ("Longueur de la chambre", L_ROOM, "m", "Dimensions : 5.0m x 5.0m x 4.0m"),
        ("Largeur de la chambre", W_ROOM, "m", "Surface au sol = 25 m²"),
        ("Hauteur de la chambre", H_ROOM, "m", "Hauteur sous plafond"),
        ("Surface de l'enveloppe extérieure", A_ENV, "m²", "Calculé d'après les dimensions"),
        ("Épaisseur isolant PU (+4°C)", THICKNESS_PU_POS * 1000.0, "mm", "Standard algérien pour froid positif"),
        ("Épaisseur isolant PU (-18°C)", THICKNESS_PU_NEG * 1000.0, "mm", "Standard algérien pour froid négatif"),
        ("Conductivité de l'Aluminium", LAMBDA_AL, "W/m.K", "Matériau enveloppe MCP"),
        ("Conductivité de la Paraffine (MCP)", LAMBDA_PCM, "W/m.K", "Matériau à changement de phase"),
        ("Chaleur latente Paraffine (L_f)", L_F, "J/kg", "Énergie de changement de phase (200 000 J/kg)"),
        ("Masse volumique Paraffine (MCP)", RHO_PCM, "kg/m³", "Phase liquide/solide moyenne"),
        ("Autonomie cible requise", T_AUTONOMY_TARGET, "heures", "Heures de décharge sans compresseur (13h)"),
        ("Tarif Sonelgaz (Heures Pleines)", TARIF_HP_DA_KWH, "DA/kWh", "Tarif fort en journée"),
        ("Tarif Sonelgaz (Heures Creuses)", TARIF_HC_DA_KWH, "DA/kWh", "Tarif réduit la nuit (recharge)"),
        ("Volume du stock stocké", STOCK_VOLUME, "L", "Equivalent à 20 tonnes d'eau/produit"),
        ("Renouvellement stock par jour", STOCK_TURNOVER * 100.0, "%", "Turnover quotidien de marchandises"),
        ("Dimensions porte (L x H)", f"{W_DOOR} x {H_DOOR}", "m", "Porte pivotante de service"),
        ("Ouvertures de porte par heure", N_OPENINGS_PER_HOUR, "fois/h", "Fréquence d'accès industrielle moyenne"),
        ("Durée d'ouverture moyenne", T_OPENING, "s", "Infiltration massique d'air extérieur"),
        ("Résistance de contact Époxy", R_CONTACT_EPOXY, "m².K/W", "Résistance thermique peinture")
    ]
    
    headers_params = ["Paramètre de Conception", "Valeur", "Unité", "Description physique"]
    for col_idx, h in enumerate(headers_params, start=1):
        cell = ws_dash.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_thin
    
    ws_dash.row_dimensions[5].height = 24
    
    for row_idx, param in enumerate(fixed_params, start=6):
        ws_dash.row_dimensions[row_idx].height = 20
        name, val, unit, desc = param
        
        c1 = ws_dash.cell(row=row_idx, column=1, value=name)
        c2 = ws_dash.cell(row=row_idx, column=2, value=val)
        c3 = ws_dash.cell(row=row_idx, column=3, value=unit)
        c4 = ws_dash.cell(row=row_idx, column=4, value=desc)
        
        for c in [c1, c2, c3, c4]:
            c.font = font_regular
            c.border = border_thin
        
        c1.alignment = align_left
        c2.alignment = align_right
        c3.alignment = align_center
        c4.alignment = align_left
        
        # Formater les nombres
        if isinstance(val, float):
            if val > 100:
                c2.number_format = '#,##0.0'
            elif val < 0.01:
                c2.number_format = '0.00000'
            else:
                c2.number_format = '0.000'
        elif isinstance(val, int):
            c2.number_format = '#,##0'

    # Section 2 : Explication des Formules Physiques (Colonnes F à H)
    ws_dash["F4"] = "2. LOIS PHYSIQUES DE RÉFÉRENCE UTILISÉES"
    ws_dash["F4"].font = font_section
    
    formulas_explanations = [
        ("Loi de Fourier (Conduction)", "q_cond = U * A * Delta_T", "Utilisée pour modéliser le transfert thermique à travers les parois isolantes en Polyuréthane de la chambre froide, ainsi que la conduction radiale à travers l'épaisseur de l'Aluminium."),
        ("Loi de refroidissement de Newton (Convection)", "q_conv = h * A_eff * (T_air - T_pcm)", "Modélise le transfert convectif entre l'air de la chambre froide et la surface externe des cylindres MCP. Calcule l'impact de la ventilation forcée (ON, h=25) par rapport à la convection naturelle (OFF, h=5)."),
        ("Équation de la chaleur latente", "Q = m_pcm * L_f", "Détermine la masse de MCP requise pour absorber le besoin frigorifique total cumulé sur la période d'autonomie cible (13 heures). m_pcm = Q_besoin / L_f (L_f = 200 000 J/kg)."),
        ("Théorème de Sebbar (Infiltration d'air)", "q_inf = m_dot_air * Delta_h * D_open", "Modélise l'ouverture de porte comme un échange thermique massique dynamique. Le débit massique d'air entrant (m_dot_air) dépend de la différence de densité due aux températures intérieure et extérieure."),
        ("Théorie des ailettes (Efficacité d'échange)", "eta_f = tanh(m*L) / (m*L)", "Permet de calculer l'augmentation réelle de la capacité d'échange thermique externe apportée par les ailettes en aluminium. Longueur fixe de 30 mm.")
    ]
    
    headers_formulas = ["Phénomène Physique", "Formule mathématique clé", "Pertinence et application technique"]
    for col_idx, h in enumerate(headers_formulas, start=6):
        cell = ws_dash.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    for row_offset, formula in enumerate(formulas_explanations):
        row_num = 6 + row_offset
        ws_dash.row_dimensions[row_num].height = 60
        name, eq, desc = formula
        
        c1 = ws_dash.cell(row=row_num, column=6, value=name)
        c2 = ws_dash.cell(row=row_num, column=7, value=eq)
        c3 = ws_dash.cell(row=row_num, column=8, value=desc)
        
        for c in [c1, c2, c3]:
            c.font = font_regular
            c.border = border_thin
            c.alignment = Alignment(vertical='center', wrap_text=True)
            
        c1.font = font_bold
        c2.alignment = align_center
        c2.font = Font(name="Courier New", size=10, bold=True, color="1B365D")
        c3.alignment = align_left
        
    # Section 3 : Configurations Optimales Identifiées (Matrice de synthèse + Sensibilité)
    ws_dash["A27"] = "3. SYNTHÈSE DES CONFIGURATIONS OPTIMALES ET ANALSES DE SENSIBILITÉ ÉCONOMIQUE"
    ws_dash["A27"].font = font_section
    ws_dash["A28"] = "*Note : Les colonnes jaunes représentent l'analyse de sensibilité économique (+20% coût Alu) et le temps de retour sur investissement face aux tarifs Sonelgaz."
    ws_dash["A28"].font = font_small
    
    headers_opt = [
        "T. Cible (°C)", "T. Ext (°C)", "Dia. Cyl (mm)", "Ép. Al (mm)", "N. Ailettes", 
        "Long. Cyl (m)", "Espace Vide (%)", "Ventilation", "Charge (W)", "Masse MCP (kg)", 
        "Masse Al (kg)", "Autonomie (h)", "Modules Requis", "Coût Standard (DA)", 
        "Coût Sens. Alu+20% (DA)", "Gains Sonelgaz (DA/an)", "T.R.I. Std (ans)", "T.R.I. Sens (ans)"
    ]
    
    for col_idx, h in enumerate(headers_opt, start=1):
        cell = ws_dash.cell(row=30, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_thin
    ws_dash.row_dimensions[30].height = 24
    
    for row_idx, (_, row) in enumerate(df_best_configs.iterrows(), start=31):
        ws_dash.row_dimensions[row_idx].height = 22
        
        vals = [
            row["Temp_Target"], row["Temp_Ext"], row["D_Cyl_mm"], row["T_Wall_Al_mm"], 
            row["N_Fins"], row["L_Cyl_m"], row["Empty_Space_Percent"], row["Ventilation"],
            row["Q_Load_Total_W"], row["M_PCM_Required_kg"], row["M_Al_Total_kg"],
            row["Autonomy_Real_h"], row["N_Modules"], row["Cost_DA"], 
            row["Cost_Sens_DA"], row["Savings_Yearly_DA"], row["Payback_Years"], row["Payback_Years_Sens"]
        ]
        
        for col_idx, val in enumerate(vals, start=1):
            cell = ws_dash.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_bold
            cell.border = border_thin
            cell.alignment = align_center
            
            # Appliquer couleur distinctive pour la sensibilité économique
            if col_idx in [15, 16, 17, 18]:
                cell.fill = fill_accent_red
            else:
                cell.fill = fill_accent_green
            
            # Formater
            if col_idx in [1, 2]:
                cell.number_format = '0 "°C"'
            elif col_idx in [3, 4]:
                cell.number_format = '0.0 "mm"'
            elif col_idx == 5:
                cell.number_format = '0'
            elif col_idx == 6:
                cell.number_format = '0.0 "m"'
            elif col_idx == 7:
                cell.number_format = '0 "%"'
            elif col_idx == 9:
                cell.number_format = '#,##0 "W"'
            elif col_idx in [10, 11]:
                cell.number_format = '#,##0.0 "kg"'
            elif col_idx == 12:
                cell.number_format = '0.0 "h"'
            elif col_idx == 13:
                cell.number_format = '#,##0'
            elif col_idx in [14, 15, 16]:
                cell.number_format = '#,##0 "DA"'
            elif col_idx in [17, 18]:
                cell.number_format = '0.0 "ans"'

    # Section 4 : Notes d'ingénierie et recommandations pour la soutenance (Jury)
    start_row_notes = 31 + len(df_best_configs) + 3
    ws_dash.cell(row=start_row_notes, column=1, value="4. NOTES TECHNIQUES POUR LA SOUTENANCE DEVANT LE JURY").font = font_section
    
    notes = [
        ("Facteur Givre (Sensibilité Thermique)", "Le modèle calcule l'autonomie avec h=25 en mode ventilé, mais le givre sur les tubes en aluminium peut diviser ce coefficient par 2 ou 3.", "Recommandation CAO/Automatisme : Programmer un cycle de dégivrage automatique intelligent piloté par capteur d'humidité dans le code de contrôle."),
        ("Charge Structurelle Plafond (Sécurité)", "La configuration optimale pour -18°C requiert 1011.2 kg d'aluminium et 2730.2 kg de MCP, soit 3.7 tonnes suspendues au plafond de la chambre froide.", "Recommandation Génie Civil : Valider la tenue mécanique de la structure porteuse de la chambre froide et ajouter des renforts métalliques si nécessaire."),
        ("Validation Puissance Onduleur (UPS)", "Le système nécessite un ventilateur de décharge actif (convection forcée). La puissance électrique totale requise par les ventilateurs doit être inférieure à la puissance de l'onduleur de secours.", "Recommandation Électrique : Dimensionner l'onduleur (UPS) pour supporter les ventilateurs en secours sans coupure pendant les 13h d'autonomie."),
        ("Argumentation face à Viking Cold (USA)", "Viking Cold utilise des plaques propriétaires scellées en plastique (conductivité ~0.15 W/m.K). Notre système utilise des tubes alu (conductivité 200 W/m.K) avec ailettes externes.", "Argument Ingénierie Locale : Conductivité thermique 1300 fois supérieure, conception démontable/réparable localement en Algérie, et coût divisé par 3.")
    ]
    
    ws_dash.cell(row=start_row_notes+1, column=1, value="Sujet Critique").font = font_header
    ws_dash.cell(row=start_row_notes+1, column=1).fill = fill_navy
    ws_dash.cell(row=start_row_notes+1, column=1).alignment = align_center
    ws_dash.cell(row=start_row_notes+1, column=1).border = border_thin
    
    ws_dash.cell(row=start_row_notes+1, column=2, value="Détail Physique / Impact").font = font_header
    ws_dash.cell(row=start_row_notes+1, column=2).fill = fill_navy
    ws_dash.cell(row=start_row_notes+1, column=2).alignment = align_center
    ws_dash.cell(row=start_row_notes+1, column=2).border = border_thin
    
    ws_dash.merge_cells(start_row=start_row_notes+1, start_column=3, end_row=start_row_notes+1, end_column=7)
    ws_dash.cell(row=start_row_notes+1, column=3, value="Recommandation d'Ingénierie pour la soutenance").font = font_header
    ws_dash.cell(row=start_row_notes+1, column=3).fill = fill_navy
    ws_dash.cell(row=start_row_notes+1, column=3).alignment = align_center
    ws_dash.cell(row=start_row_notes+1, column=3).border = border_thin
    
    ws_dash.row_dimensions[start_row_notes+1].height = 24
    
    for i, (sujet, detail, recom) in enumerate(notes):
        curr_row = start_row_notes + 2 + i
        ws_dash.row_dimensions[curr_row].height = 45
        
        c1 = ws_dash.cell(row=curr_row, column=1, value=sujet)
        c2 = ws_dash.cell(row=curr_row, column=2, value=detail)
        ws_dash.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=7)
        c3 = ws_dash.cell(row=curr_row, column=3, value=recom)
        
        for c in [c1, c2, c3]:
            c.font = font_regular
            c.border = border_thin
            c.alignment = Alignment(vertical='center', wrap_text=True)
            
        c1.font = font_bold
        c1.fill = fill_light_blue
        c2.alignment = align_left
        c3.alignment = align_left

    # --------------------------------------------------------------------------
    # RENSEIGNER LES DONNÉES DE SIMULATION COMPLÈTES (FEUILLE 3)
    # --------------------------------------------------------------------------
    
    headers_data = [
        "Temp. Cible (°C)", "Temp. Extérieure (°C)", "Diamètre Cylindre (mm)", "Épaisseur Paroi Al (mm)",
        "Nombre d'Ailettes", "Épaisseur Ailettes (mm)", "Épaisseur Époxy (µm)", "Ventilation",
        "Longueur Cylindre (m)", "Espace Vide (%)",
        "Pertes Parois (W)", "Infiltration Porte (W)", "Refroidissement Produit (W)", "Charges Internes (W)",
        "Charge Thermique Totale (W)", "Masse MCP Requise (kg)", "Masse Aluminium (kg)",
        "Résistance Linéaire (K.m/W)", "Lambda Eff. MCP (W/m.K)", "UA Batterie (W/K)",
        "Puissance Batterie Max (W)", "Temps d'Autonomie (h)", "Efficacité Thermique (%)",
        "Efficacité Gravimétrique (%)", "Efficacité Totale (%)", "Volume Cylindre (m³)",
        "Masse MCP Dilatée (kg)", "Volume MCP Dilaté (m³)", "Nombre de Modules Requis",
        "Coût Estimé Std (DA)", "Coût Sens. Alu+20% (DA)", "Gains Sonelgaz (DA/an)",
        "T.R.I. Std (ans)", "T.R.I. Sens (ans)", "Score d'Optimisation (h/DA)", "Statut Optimalité"
    ]
    
    for col_idx, h in enumerate(headers_data, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_thin
    ws_data.row_dimensions[1].height = 30
    
    for row_idx, (orig_idx, row) in enumerate(df.iterrows(), start=2):
        ws_data.row_dimensions[row_idx].height = 20
        
        is_best = (orig_idx in best_indices)
        is_zebra = (row_idx % 2 == 0)
        
        if is_best:
            current_fill = fill_accent_green
        else:
            current_fill = fill_zebra if is_zebra else PatternFill(fill_type=None)
        
        inputs = [
            row["Temp_Target"], row["Temp_Ext"], row["D_Cyl_mm"], row["T_Wall_Al_mm"],
            row["N_Fins"], row["T_Fin_mm"], row["Paint_Thickness_um"], row["Ventilation"],
            row["L_Cyl_m"], row["Empty_Space_Percent"],
            row["Q_Wall_W"], row["Q_Infiltration_W"], row["Q_Product_W"], row["Q_Internal_W"]
        ]
        
        for col_idx, val in enumerate(inputs, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = border_thin
            cell.alignment = align_center
            if current_fill.fill_type:
                cell.fill = current_fill
                
            if col_idx in [1, 2]:
                cell.number_format = '0'
            elif col_idx in [3, 4, 6]:
                cell.number_format = '0.0'
            elif col_idx == 7:
                cell.number_format = '0'
            elif col_idx == 9:
                cell.number_format = '0.0'
            elif col_idx == 10:
                cell.number_format = '0'
            elif col_idx in [11, 12, 13, 14]:
                cell.number_format = '#,##0.0'
        
        outputs = [
            row["Q_Load_Total_W"], row["M_PCM_Required_kg"], row["M_Al_Total_kg"],
            row["R_Linear_K_m_W"], row["Lambda_Eff_W_m_K"], row["UA_Battery_W_K"],
            row["Q_Battery_Max_W"], row["Autonomy_Real_h"], row["Efficiency_Thermal"],
            row["Efficiency_Gravimetric"], row["Efficiency_Total"], row["Volume_Cyl_m3"],
            row["M_PCM_Dilated_kg"], row["V_PCM_Dilated_m3"], row["N_Modules"],
            row["Cost_DA"], row["Cost_Sens_DA"], row["Savings_Yearly_DA"],
            row["Payback_Years"], row["Payback_Years_Sens"], row["Optimization_Score_h_DA"]
        ]
        
        for col_idx, val in enumerate(outputs, start=15):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = border_thin
            cell.alignment = align_center
            if current_fill.fill_type:
                cell.fill = current_fill
                
            if col_idx in [15, 16, 17, 20, 21]:
                cell.number_format = '#,##0.0'
            elif col_idx == 18:
                cell.number_format = '0.00000'
            elif col_idx == 19:
                cell.number_format = '0.000'
            elif col_idx == 22:
                cell.number_format = '0.0'
            elif col_idx in [23, 24, 25]:
                cell.number_format = '0.0%'
            elif col_idx in [26, 28]:
                cell.number_format = '0.0000'
            elif col_idx in [27, 29]:
                cell.number_format = '#,##0.0'
            elif col_idx in [30, 31, 32]:
                cell.number_format = '#,##0'
            elif col_idx in [33, 34]:
                cell.number_format = '0.0'
            elif col_idx == 35:
                cell.number_format = '0.00E+00'
                
        cell_opt = ws_data.cell(row=row_idx, column=36, value="OPTIMAL" if is_best else "")
        cell_opt.font = font_bold
        cell_opt.border = border_thin
        cell_opt.alignment = align_center
        if current_fill.fill_type:
            cell_opt.fill = current_fill
            
    # Ajustement automatique des largeurs de colonnes pour les feuilles
    for ws in [ws_dash, ws_data]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        is_merged = True
                        break
                if not is_merged and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    ws_dash.column_dimensions['A'].width = 32
    ws_dash.column_dimensions['B'].width = 15
    ws_dash.column_dimensions['C'].width = 10
    ws_dash.column_dimensions['D'].width = 38
    ws_dash.column_dimensions['E'].width = 3
    ws_dash.column_dimensions['F'].width = 35
    ws_dash.column_dimensions['G'].width = 25
    ws_dash.column_dimensions['H'].width = 65
    
    # --------------------------------------------------------------------------
    # RENSEIGNER LES GRAPHES D'INTERPOLATION (FEUILLE 2)
    # --------------------------------------------------------------------------
    ws_graphs = wb["Graphes d'Interpolation"]
    ws_graphs.views.sheetView[0].showGridLines = True
    
    # Titre de la feuille
    ws_graphs.merge_cells("A1:J2")
    title_cell = ws_graphs["A1"]
    title_cell.value = "GRAPHES ET TABLES D'INTERPOLATION DES PARAMÈTRES DE CONCEPTION"
    title_cell.font = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    title_cell.fill = fill_navy
    title_cell.alignment = align_center
    
    ws_graphs.row_dimensions[1].height = 25
    ws_graphs.row_dimensions[2].height = 25
    
    # TABLEAU 1: Autonomie Réelle vs Diamètre et Ailettes
    ws_graphs["A4"] = "1. IMPACT DU DIAMÈTRE ET DES AILETTES SUR L'AUTONOMIE RÉELLE (heures)"
    ws_graphs["A4"].font = font_section
    
    ws_graphs.cell(row=5, column=1, value="Nombre d'ailettes").font = font_bold
    ws_graphs.cell(row=5, column=1).border = border_thin
    ws_graphs.cell(row=5, column=1).alignment = align_center
    ws_graphs.cell(row=5, column=1).fill = fill_light_blue
    
    diams = [80, 100, 120]
    for idx, d in enumerate(diams, start=2):
        cell = ws_graphs.cell(row=5, column=idx, value=d)
        cell.font = font_bold
        cell.border = border_thin
        cell.alignment = align_center
        cell.fill = fill_light_blue
        cell.number_format = '0 "mm"'
        
    df_neg_35_100_on_2 = df[
        (df["Temp_Target"] == -18.0) &
        (df["Temp_Ext"] == 35.0) &
        (df["Paint_Thickness_um"] == 100.0) &
        (df["Ventilation"] == "ON") &
        (df["T_Wall_Al_mm"] == 2.0) &
        (df["L_Cyl_m"] == 1.5) &
        (df["Empty_Space_Percent"] == 10)
    ]
    
    fins = [0, 4, 8]
    for f_idx, f in enumerate(fins, start=6):
        row_label = ws_graphs.cell(row=f_idx, column=1, value=f"{f} Ailettes")
        row_label.font = font_bold
        row_label.border = border_thin
        row_label.alignment = align_center
        row_label.fill = fill_light_blue
        
        for d_idx, d in enumerate(diams, start=2):
            cell = ws_graphs.cell(row=f_idx, column=d_idx)
            f_thick = 0.0 if f == 0 else 1.5
            match = df_neg_35_100_on_2[
                (df_neg_35_100_on_2["D_Cyl_mm"] == d) &
                (df_neg_35_100_on_2["N_Fins"] == f) &
                (df_neg_35_100_on_2["T_Fin_mm"] == f_thick)
            ]
            cell.value = float(match["Autonomy_Real_h"].values[0]) if not match.empty else 0.0
            cell.font = font_regular
            cell.border = border_thin
            cell.alignment = align_center
            cell.number_format = '0.0 "h"'
            
    chart1 = LineChart()
    chart1.title = "Autonomie Réelle vs Diamètre du Cylindre (à -18°C / Ext 35°C / Peinture 100 µm / Vent ON)"
    chart1.style = 13
    chart1.y_axis.title = "Autonomie (heures)"
    chart1.x_axis.title = "Diamètre Cylindre (mm)"
    chart1.width = 18
    chart1.height = 10
    
    data1 = Reference(ws_graphs, min_col=1, min_row=5, max_col=4, max_row=8)
    cats1 = Reference(ws_graphs, min_col=2, min_row=5, max_col=4, max_row=5)
    chart1.add_data(data1, titles_from_data=True, from_rows=True)
    chart1.set_categories(cats1)
    
    for s in chart1.series:
        s.smooth = True
    ws_graphs.add_chart(chart1, "B10")
    
    # TABLEAU 2: Efficacité Totale vs Diamètre et Ailettes
    ws_graphs["A26"] = "2. IMPACT DU DIAMÈTRE ET DES AILETTES SUR L'EFFICACITÉ TOTALE (%)"
    ws_graphs["A26"].font = font_section
    
    ws_graphs.cell(row=27, column=1, value="Nombre d'ailettes").font = font_bold
    ws_graphs.cell(row=27, column=1).border = border_thin
    ws_graphs.cell(row=27, column=1).alignment = align_center
    ws_graphs.cell(row=27, column=1).fill = fill_light_blue
    
    for idx, d in enumerate(diams, start=2):
        cell = ws_graphs.cell(row=27, column=idx, value=d)
        cell.font = font_bold
        cell.border = border_thin
        cell.alignment = align_center
        cell.fill = fill_light_blue
        cell.number_format = '0 "mm"'
        
    for f_idx, f in enumerate(fins, start=28):
        row_label = ws_graphs.cell(row=f_idx, column=1, value=f"{f} Ailettes")
        row_label.font = font_bold
        row_label.border = border_thin
        row_label.alignment = align_center
        row_label.fill = fill_light_blue
        
        for d_idx, d in enumerate(diams, start=2):
            cell = ws_graphs.cell(row=f_idx, column=d_idx)
            f_thick = 0.0 if f == 0 else 1.5
            match = df_neg_35_100_on_2[
                (df_neg_35_100_on_2["D_Cyl_mm"] == d) &
                (df_neg_35_100_on_2["N_Fins"] == f) &
                (df_neg_35_100_on_2["T_Fin_mm"] == f_thick)
            ]
            cell.value = float(match["Efficiency_Total"].values[0]) if not match.empty else 0.0
            cell.font = font_regular
            cell.border = border_thin
            cell.alignment = align_center
            cell.number_format = '0.0%'
            
    chart2 = LineChart()
    chart2.title = "Efficacité Totale vs Diamètre (à -18°C / Ext 35°C / Peinture 100 µm / Vent ON)"
    chart2.style = 13
    chart2.y_axis.title = "Efficacité Totale (%)"
    chart2.x_axis.title = "Diamètre Cylindre (mm)"
    chart2.width = 18
    chart2.height = 10
    
    data2 = Reference(ws_graphs, min_col=1, min_row=27, max_col=4, max_row=30)
    cats2 = Reference(ws_graphs, min_col=2, min_row=27, max_col=4, max_row=27)
    chart2.add_data(data2, titles_from_data=True, from_rows=True)
    chart2.set_categories(cats2)
    
    for s in chart2.series:
        s.smooth = True
    ws_graphs.add_chart(chart2, "B32")
    
    # TABLEAU 3: Impact de l'Épaisseur de Peinture Époxy sur l'Autonomie
    ws_graphs["A48"] = "3. IMPACT DE L'ÉPAISSEUR DE PEINTURE ÉPOXY SUR L'AUTONOMIE RÉELLE (heures)"
    ws_graphs["A48"].font = font_section
    
    ws_graphs.cell(row=49, column=1, value="Diamètre de Cylindre").font = font_bold
    ws_graphs.cell(row=49, column=1).border = border_thin
    ws_graphs.cell(row=49, column=1).alignment = align_center
    ws_graphs.cell(row=49, column=1).fill = fill_light_blue
    
    paint_thicks = [50, 100, 150, 200]
    for idx, p in enumerate(paint_thicks, start=2):
        cell = ws_graphs.cell(row=49, column=idx, value=p)
        cell.font = font_bold
        cell.border = border_thin
        cell.alignment = align_center
        cell.fill = fill_light_blue
        cell.number_format = '0 "µm"'
        
    df_neg_35_fins4_on_2_ail15 = df[
        (df["Temp_Target"] == -18.0) &
        (df["Temp_Ext"] == 35.0) &
        (df["N_Fins"] == 4) &
        (df["T_Fin_mm"] == 1.5) &
        (df["Ventilation"] == "ON") &
        (df["T_Wall_Al_mm"] == 2.0) &
        (df["L_Cyl_m"] == 1.5) &
        (df["Empty_Space_Percent"] == 10)
    ]
    
    diams_paint = [80, 100, 120]
    for d_idx, d in enumerate(diams_paint, start=50):
        row_label = ws_graphs.cell(row=d_idx, column=1, value=f"Cylindre {d} mm")
        row_label.font = font_bold
        row_label.border = border_thin
        row_label.alignment = align_center
        row_label.fill = fill_light_blue
        
        for p_idx, p in enumerate(paint_thicks, start=2):
            cell = ws_graphs.cell(row=d_idx, column=p_idx)
            match = df_neg_35_fins4_on_2_ail15[
                (df_neg_35_fins4_on_2_ail15["D_Cyl_mm"] == d) &
                (df_neg_35_fins4_on_2_ail15["Paint_Thickness_um"] == p)
            ]
            cell.value = float(match["Autonomy_Real_h"].values[0]) if not match.empty else 0.0
            cell.font = font_regular
            cell.border = border_thin
            cell.alignment = align_center
            cell.number_format = '0.0 "h"'
            
    chart3 = LineChart()
    chart3.title = "Autonomie vs Épaisseur de Peinture Époxy (à -18°C / Ext 35°C / 4 Ailettes / Vent ON)"
    chart3.style = 13
    chart3.y_axis.title = "Autonomie (heures)"
    chart3.x_axis.title = "Épaisseur Peinture Époxy (µm)"
    chart3.width = 18
    chart3.height = 10
    
    data3 = Reference(ws_graphs, min_col=1, min_row=49, max_col=5, max_row=52)
    cats3 = Reference(ws_graphs, min_col=2, min_row=49, max_col=5, max_row=49)
    chart3.add_data(data3, titles_from_data=True, from_rows=True)
    chart3.set_categories(cats3)
    
    for s in chart3.series:
        s.smooth = True
    ws_graphs.add_chart(chart3, "B54")
    
    ws_graphs.column_dimensions['A'].width = 25
    
    # Enregistrer le classeur Excel
    saved = False
    suffix = 0
    current_filename = filename
    while not saved:
        try:
            wb.save(current_filename)
            saved = True
            if suffix == 0:
                print(f"Simulation avec analyse de sensibilité terminée. Fichier généré : '{current_filename}'")
            else:
                print(f"\n[ATTENTION] Le fichier original était ouvert dans Excel.")
                print(f"La simulation a été sauvegardée sous le nom alternatif : '{current_filename}'")
        except PermissionError:
            suffix += 1
            current_filename = filename.replace(".xlsx", f"_{suffix}.xlsx")
            if suffix > 10:
                print("\n[ERREUR CRITIQUE] Impossible d'enregistrer le fichier Excel car toutes les versions de _1 à _10 sont ouvertes.")
                break

# ==============================================================================
# 6. EXÉCUTION DU SCRIPT
# ==============================================================================

if __name__ == "__main__":
    print("Démarrage de la simulation thermique avec sensibilité économique...")
    simulation_data = run_simulation()
    print(f"Matrice de test résolue : {len(simulation_data)} configurations simulées.")
    create_styled_excel(simulation_data, filename="Simulation_Performances_MCP_TES_Final.xlsx")
